"""
=================================
Author: Flora Chen
Time: 2021/1/20 20:34
-_- -_- -_- -_- -_- -_- -_- -_-
=================================
"""

"""
打开excel --> 选择sheet表格  --> 获取数据
"""

import openpyxl, os


class ExcelHandler:
    def __init__(self, path):
        self.path = path

    def read(self, sheet_name):
        """
        读取数据
        :param sheet_name: 表单名称
        :return: data: 测试数据 字典格式
        """
        # 打开excel
        workbook = openpyxl.open(self.path)

        # 选择sheet
        worksheet = workbook[sheet_name]

        # 关闭文件
        workbook.close()

        all_values = list(worksheet.values)

        header = all_values[0]

        data = []
        for row in all_values[1:]:
            data.append(dict(zip(header, row)))

        return data

    def write(self, sheetname, row, column, data):
        """
        往excel中写入数据
        :param sheetname:  表单名称
        :return: None
        """
        # 打开excel
        workbook = openpyxl.load_workbook(self.path)

        # 选择sheet
        worksheet = workbook[sheetname]

        # 写入数据
        worksheet.cell(row=row, column=column).value = data

        # 保存并关闭文件
        workbook.save(self.path)
        workbook.close()


if __name__ == "__main__":
    from conf.path import DATA_DIR

    # 获取excel的路径
    excel_file = os.path.join(DATA_DIR, "case.xlsx")
    # 获取excel中的测试数据
    data = ExcelHandler(excel_file).read("register")
    print(data)


"""

# 打开excel
workbook = openpyxl.open(os.path.join(CASE_DIR, "case.xlsx"))

# 选择sheet
sheet= workbook["register"]

# 获取某个单元格数据对象
cell = sheet.cell(row=1, column=2)

# 获取单元格数据对象的值
print(cell.value)

# 获取某一行数据,输出的是数据对象
print(sheet[2])

# 获取所有行,输出的是数据对象
print(list(sheet.rows))

# 获取所有数据
print(list(sheet.values))
"""
